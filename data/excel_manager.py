from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from data.models import Game
from config import EXCEL_PATH, PRIORITY_COLORS
import data.repository as repo


# ── Color constants ────────────────────────────────────────────────
BG_HEADER   = "1B2838"
BG_ROW_ODD  = "1E2A38"
BG_ROW_EVEN = "17202A"
FG_TEXT     = "C7D5E0"
FG_DIM      = "8F98A0"
FG_BLUE     = "66C0F4"
FG_GREEN    = "5BA32B"
FG_GOLD     = "E4A81E"
FG_RED      = "C94444"

PRIORITY_BG = {"S": "E4A81E", "A": "5BA32B", "B": "3D7EAA", "C": "666666"}

THIN = Side(style="thin", color="2A3F5F")
BORDER = Border(bottom=Side(style="thin", color="2A3F5F"))

COLUMNS = [
    ("ID",                  5),
    ("Nombre",              30),
    ("Steam AppID",         12),
    ("Steam URL",           35),
    ("Género",              18),
    ("Año",                  6),
    ("Desarrollador",       22),
    ("Publisher",           22),
    ("Categorías",          25),
    ("Prioridad",            9),
    ("Estado",              10),
    ("Precio actual",       14),
    ("Moneda",               8),
    ("Descuento %",          11),
    ("Precio base",         12),
    ("Mínimo histórico",    16),
    ("Fecha mínimo",        14),
    ("Diferencia %",        12),
    ("Valoración",          10),
    ("Notas",               30),
    ("Fecha agregado",      14),
    ("Portada",             16),
]


def _header_font(bold=True):
    return Font(name="Arial", bold=bold, color=FG_TEXT, size=9)


def _cell_font(color=FG_TEXT, bold=False, size=9):
    return Font(name="Arial", color=color, bold=bold, size=size)


def _fill(hex_color: str):
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)


def export_excel() -> Path:
    """Generate (or regenerate) the full Excel workbook from the repository."""
    games = repo.get_all()

    wb = Workbook()
    _build_wishlist_sheet(wb, games)
    _build_dashboard_sheet(wb, games)

    # Remove default sheet if it still exists
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    wb.save(EXCEL_PATH)
    return EXCEL_PATH


def _build_wishlist_sheet(wb: Workbook, games: list[Game]) -> None:
    ws = wb.create_sheet("Wishlist")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    # Header row
    for col_idx, (col_name, col_width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = _header_font()
        cell.fill = _fill(BG_HEADER)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    ws.row_dimensions[1].height = 22

    # Data rows
    for row_idx, game in enumerate(games, start=2):
        bg = BG_ROW_ODD if row_idx % 2 == 0 else BG_ROW_EVEN
        row_fill = _fill(bg)

        price = game.price
        history = game.price_history

        values = [
            game.id,
            game.name,
            game.app_id,
            game.steam_url,
            game.genre,
            game.release_year or "",
            game.developer,
            game.publisher,
            game.categories,
            game.priority,
            game.status,
            price.current if price else "",
            price.currency if price else "",
            f"-{price.discount_pct}%" if price and price.discount_pct > 0 else "",
            price.base if price else "",
            history.all_time_low if history else "",
            history.all_time_low_date if history else "",
            f"{game.price_diff_pct:.1f}%" if game.price_diff_pct is not None else "",
            game.personal_rating or "",
            game.notes,
            game.date_added,
            "",  # Cover image column
        ]

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = row_fill
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=False)

            # Column-specific styling
            col_name = COLUMNS[col_idx - 1][0]

            if col_name == "Prioridad":
                p_color = PRIORITY_BG.get(str(value), "666666")
                cell.fill = _fill(p_color)
                cell.font = _cell_font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")

            elif col_name == "Nombre":
                cell.font = _cell_font(color=FG_TEXT, bold=True)

            elif col_name == "Steam URL":
                cell.font = Font(name="Arial", color=FG_BLUE, size=9, underline="single")
                cell.hyperlink = str(value) if value else None

            elif col_name == "Precio actual" and price and price.is_on_sale:
                cell.font = _cell_font(color=FG_GREEN, bold=True)

            elif col_name == "Diferencia %":
                diff = game.price_diff_pct
                if diff is not None:
                    color = FG_GREEN if diff <= 5 else (FG_GOLD if diff <= 20 else FG_RED)
                    cell.font = _cell_font(color=color, bold=True)
                else:
                    cell.font = _cell_font(color=FG_DIM)

            elif col_name == "Mínimo histórico":
                cell.font = _cell_font(color=FG_BLUE)

            else:
                cell.font = _cell_font()

        ws.row_dimensions[row_idx].height = 20

        # Embed cover image if available
        if game.cover_path and Path(game.cover_path).exists():
            try:
                img = XLImage(game.cover_path)
                img.width = 40
                img.height = 53
                cover_col = get_column_letter(len(COLUMNS))
                ws.row_dimensions[row_idx].height = 55
                ws.add_image(img, f"{cover_col}{row_idx}")
            except Exception:
                pass


def _build_dashboard_sheet(wb: Workbook, games: list[Game]) -> None:
    ws = wb.create_sheet("Dashboard")
    ws.sheet_view.showGridLines = False

    title_font = Font(name="Arial", bold=True, color=FG_BLUE, size=14)
    label_font = Font(name="Arial", color=FG_DIM, size=9)
    value_font = Font(name="Arial", bold=True, color=FG_TEXT, size=11)
    header_fill = _fill(BG_HEADER)
    card_fill = _fill("1E2A38")

    ws.column_dimensions["A"].width = 3
    for col in ["B", "C", "D", "E", "F", "G", "H"]:
        ws.column_dimensions[col].width = 18

    # Title
    ws["B2"] = "📊  Dashboard — Steam Library Curator"
    ws["B2"].font = title_font
    ws["B2"].fill = header_fill
    ws.merge_cells("B2:H2")
    ws["B2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 30

    # ── Stat cards (row 4) ────────────────────────────────────────
    stat_row = 4
    ws.row_dimensions[stat_row].height = 14
    ws.row_dimensions[stat_row + 1].height = 18

    total = len(games)
    on_sale = sum(1 for g in games if g.price and g.price.is_on_sale)
    total_value = sum(g.price.current for g in games if g.price)
    min_value = sum(
        g.price_history.all_time_low for g in games
        if g.price_history and g.price_history.all_time_low > 0
    )
    savings = max(0, total_value - min_value)
    priority_s = sum(1 for g in games if g.priority == "S")

    stats = [
        ("Juegos en wishlist", total, FG_BLUE),
        ("En oferta ahora",    on_sale, FG_GREEN),
        ("Valor actual (MXN)", f"${total_value:,.0f}", FG_TEXT),
        ("Al mínimo hist.",    f"${min_value:,.0f}", FG_BLUE),
        ("Ahorro potencial",   f"${savings:,.0f}", FG_GREEN),
        ("Prioridad S",        priority_s, FG_GOLD),
    ]

    for i, (label, value, color) in enumerate(stats):
        col = get_column_letter(2 + i)
        label_cell = ws[f"{col}{stat_row}"]
        label_cell.value = label
        label_cell.font = label_font
        label_cell.fill = card_fill
        label_cell.alignment = Alignment(horizontal="center")

        val_cell = ws[f"{col}{stat_row + 1}"]
        val_cell.value = value
        val_cell.font = Font(name="Arial", bold=True, color=color, size=13)
        val_cell.fill = card_fill
        val_cell.alignment = Alignment(horizontal="center")

    # ── Priority table (row 7) ────────────────────────────────────
    _section_header(ws, "B7", "Distribución por Prioridad")
    ws.merge_cells("B7:C7")
    priorities = {"S": 0, "A": 0, "B": 0, "C": 0}
    for g in games:
        if g.priority in priorities:
            priorities[g.priority] += 1

    for i, (p, count) in enumerate(priorities.items()):
        r = 8 + i
        pct = f"{count/total*100:.1f}%" if total > 0 else "0%"
        _table_row(ws, r, [p, count, pct], [
            Font(name="Arial", bold=True, color="FFFFFF", size=9),
            value_font,
            Font(name="Arial", color=FG_DIM, size=9),
        ], fills=[_fill(PRIORITY_BG.get(p, "666666")), card_fill, card_fill])

    # ── Genre table (row 14) ──────────────────────────────────────
    _section_header(ws, "E7", "Distribución por Género")
    ws.merge_cells("E7:G7")
    genres: dict[str, int] = {}
    for g in games:
        for genre in g.genre.split(","):
            genre = genre.strip()
            if genre:
                genres[genre] = genres.get(genre, 0) + 1
    genres = dict(sorted(genres.items(), key=lambda x: x[1], reverse=True)[:8])

    for i, (genre, count) in enumerate(genres.items()):
        r = 8 + i
        pct = f"{count/total*100:.1f}%" if total > 0 else "0%"
        _table_row(ws, r, [genre, count, pct], [
            Font(name="Arial", color=FG_TEXT, size=9),
            value_font,
            Font(name="Arial", color=FG_DIM, size=9),
        ], start_col=5)

    # ── Decade table (row 7, col H) ───────────────────────────────
    _section_header(ws, "H7", "Por Décadas")
    decades: dict[str, int] = {}
    for g in games:
        if g.release_year:
            decade = f"{(g.release_year // 10) * 10}s"
            decades[decade] = decades.get(decade, 0) + 1
    decades = dict(sorted(decades.items()))

    for i, (decade, count) in enumerate(decades.items()):
        r = 8 + i
        _table_row(ws, r, [decade, count], [
            Font(name="Arial", color=FG_TEXT, size=9),
            value_font,
        ], start_col=8)

    # Fill remaining cells with dark background
    for row in ws.iter_rows(min_row=1, max_row=30, min_col=1, max_col=10):
        for cell in row:
            if cell.fill.fgColor.rgb in ("00000000", "FFFFFFFF") or not cell.fill.fgColor.rgb:
                cell.fill = _fill("171A21")


def _section_header(ws, cell_ref: str, title: str) -> None:
    cell = ws[cell_ref]
    cell.value = title
    cell.font = Font(name="Arial", bold=True, color="66C0F4", size=10)
    cell.fill = PatternFill("solid", start_color="1B2838", fgColor="1B2838")
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[int(cell_ref[1:])].height = 18


def _table_row(ws, row: int, values: list, fonts: list, fills: list = None, start_col: int = 2) -> None:
    for i, (val, font) in enumerate(zip(values, fonts)):
        col = get_column_letter(start_col + i)
        cell = ws[f"{col}{row}"]
        cell.value = val
        cell.font = font
        cell.fill = fills[i] if fills else PatternFill("solid", start_color="1E2A38", fgColor="1E2A38")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=Side(style="thin", color="2A3F5F"))
        ws.row_dimensions[row].height = 16
